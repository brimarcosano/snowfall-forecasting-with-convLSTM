import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta

def process_rh_file():
    # Specify the path to your Excel file
    excel_file = '/Users/briannamarcosano/Documents/SF_Testing_code/CNN_LSTM_Thesis/LSTM/Data/RelativeHumidity/RH_Merrimack.xlsx'

    # Load the workbook
    wb = load_workbook(filename=excel_file, read_only=True, data_only=True)

    # Get the active sheet
    sheet = wb.active

    # Read the data into a list of lists, skipping the first row
    data = [list(row) for row in sheet.iter_rows(min_row=2, values_only=True)]

    # Create DataFrame from the parsed data
    df = pd.DataFrame(data, columns=['INCORRECT_DATE', 'hour','County', 'RH'])

    # Ensure 'hour' and 'RH' are numeric
    df['hour'] = pd.to_numeric(df['hour'], errors='coerce')
    df['RH'] = pd.to_numeric(df['RH'], errors='coerce')

    # Generate correct dates
    start_date = datetime(1998, 1, 1)
    df['DATE'] = [start_date + timedelta(days=i//4) for i in range(len(df))]
    df['DATE'] = df['DATE'].dt.strftime('%Y/%m/%d')

    # Convert DATETIME to datetime type and extract the correct date
    df['DATE'] = pd.to_datetime(df['DATE']).dt.strftime('%Y/%m/%d')

    # Ensure 'RH' is numeric
    df['RH'] = pd.to_numeric(df['RH'], errors='coerce')

    # # Group by DATE and calculate the average RH
    # df_averaged = df.groupby(['DATE','County'])['RH'].mean().reset_index()

    # # Round the averaged RH to 2 decimal places
    # df_averaged['RH'] = df_averaged['RH'].round(2)

    return df



